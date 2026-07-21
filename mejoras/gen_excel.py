"""Genera incidencias.xlsx a partir de incidencias.json (la fuente de verdad).

Uso:
    python gen_excel.py
Requiere: openpyxl
"""
import json
import os
from datetime import datetime, timezone, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

AQUI = os.path.dirname(os.path.abspath(__file__))
DATA = json.load(open(os.path.join(AQUI, "incidencias.json"), encoding="utf-8"))


def _lima(iso_utc: str) -> str:
    """Convierte un timestamp ISO (UTC) a texto legible en hora de Perú."""
    if not iso_utc:
        return "—"
    try:
        dt = datetime.fromisoformat(iso_utc)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        dt = dt.astimezone(timezone(timedelta(hours=-5)))
        return dt.strftime("%d/%m/%Y %H:%M") + " (hora Perú)"
    except Exception:
        return iso_utc

ESTADO_FILL = {
    "Pendiente": "FFF2CC",   # ámbar
    "Aplicado":  "D8EEDC",   # verde
    "Rechazado": "F8D7DA",   # rojo
}
ESTADO_TXT = {
    "Pendiente": "9C6500",
    "Aplicado":  "1E7B34",
    "Rechazado": "A61B29",
}

wb = Workbook()
ws = wb.active
ws.title = "Incidencias"

thin = Side(style="thin", color="D0D5DB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(vertical="top", wrap_text=True)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Título
ws.merge_cells("A1:G1")
ws["A1"] = DATA["titulo"]
ws["A1"].font = Font(bold=True, size=14, color="16233A")
ws.merge_cells("A2:G2")
ws["A2"] = (f"Actualizado: {DATA['actualizado']}  ·  Última revisión de chats: "
            f"{_lima(DATA.get('ultima_revision',''))}  ·  Cada mejora requiere aprobación antes de aplicarse.")
ws["A2"].font = Font(italic=True, size=10, color="64748B")

# Encabezado
headers = ["#", "Área", "Incidencia (qué falla)", "Mejora propuesta",
           "Resumen del cambio (en simple)", "Estado", "Motivo (si rechazado)"]
ws.append([])  # fila 3 vacía
ws.append(headers)  # fila 4
hr = ws.max_row
for c in range(1, 8):
    cell = ws.cell(hr, c)
    cell.fill = PatternFill("solid", fgColor="16233A")
    cell.font = Font(bold=True, color="FFFFFF")
    cell.alignment = center
    cell.border = border

for it in DATA["incidencias"]:
    ws.append([it["id"], it["area"], it["incidencia"], it["mejora"], it["resumen"],
               it["estado"], it.get("motivo", "")])
    r = ws.max_row
    for c in range(1, 8):
        cell = ws.cell(r, c)
        cell.border = border
        cell.alignment = center if c in (1, 6) else wrap
    est = it["estado"]
    ec = ws.cell(r, 6)
    ec.fill = PatternFill("solid", fgColor=ESTADO_FILL.get(est, "FFFFFF"))
    ec.font = Font(bold=True, color=ESTADO_TXT.get(est, "000000"))

# Anchos
for col, w in zip("ABCDEFG", [5, 20, 44, 44, 44, 13, 40]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A5"
ws.auto_filter.ref = f"A4:G{ws.max_row}"

out = os.path.join(AQUI, "incidencias.xlsx")
wb.save(out)
print("Generado:", out, f"({len(DATA['incidencias'])} incidencias)")
